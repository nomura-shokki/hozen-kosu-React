from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.urls import reverse_lazy
from django.utils.timezone import make_aware
from django.views.generic import TemplateView, ListView
from django.views.generic.edit import FormView, DeleteView
from django.http import JsonResponse, HttpResponseRedirect
from pathlib import Path
import openpyxl
import datetime
import math
import os
import sys
import logging
import environ
from ..utils.kosu_utils import get_member
from ..utils.main_utils import has_non_halfwidth_characters
from ..utils.main_utils import history_record
from ..models import member, Business_Time_graph, kosu_division, team_member, administrator_data, Operation_history
from ..forms import loginForm, administrator_data_Form, uploadForm, history_findForm





#--------------------------------------------------------------------------------------------------------





# ログイン画面定義
class LoginView(FormView):
  # テンプレート,フォーム,リダイレクト先定義
  template_name = 'kosu/login.html'
  form_class = loginForm
  success_url = '/'


  # リクエストを処理するメソッドをオーバーライド
  def dispatch(self, request, *args, **kwargs):
    # ログイン済みならメイン画面にリダイレクト
    if request.session.get('login_No') and request.session.get('input_def'):
      return redirect(self.success_url)
    return super().dispatch(request, *args, **kwargs)

  # フォームが有効な場合の処理
  def form_valid(self, form):
    # フォームから送信された値を取得
    find = form.cleaned_data['employee_no4']
    # 従業員番号が空の場合、リダイレクト
    if find in ["", None]:
      messages.error(self.request, '従業員番号を入力して下さい。ERROR047')
      return redirect('/login')

    # 人員登録データから値を検索
    data = member.objects.filter(employee_no=find)
    # 従業員番号の登録がない場合、リダイレクト
    if not data.exists():
      messages.error(self.request, '入力された従業員番号は登録がありません。ERROR048')
      return redirect('/login')

    # 従業員番号の登録がある場合の処理
    else:
      # 従業員番号をセッションに保存
      self.request.session['login_No'] = find
      # 工数区分を読み込む（最新のもの）
      def_Ver = kosu_division.objects.order_by("id").last()

      # 工数区分定義が存在しない場合、リダイレクト
      if not def_Ver:
        messages.error(self.request, '利用可能な工数区分がありません。ERROR052')
        return redirect('/login')
      else:
        self.request.session['input_def'] = def_Ver.kosu_name

      # メインページにリダイレクト
      return redirect(self.success_url)


  # フォームが無効だった場合の処理（同じ画面に留まらせ）
  def form_invalid(self, form):
    return self.render_to_response(self.get_context_data(form=form))


  # コンテキストデータを設定
  def get_context_data(self, **kwargs):
    context = super().get_context_data(**kwargs)
    context['title'] = 'ログイン'
    return context





#--------------------------------------------------------------------------------------------------------





# メイン画面定義
class MainView(TemplateView):
  # テンプレート定義
  template_name = 'kosu/main.html'


  # リクエストを処理するメソッドをオーバーライド
  def dispatch(self, request, *args, **kwargs):
    # 人員情報取得
    member_obj = get_member(request)
    # 人員情報なしor未ログインの場合ログイン画面へ
    if isinstance(member_obj, HttpResponseRedirect):
      return member_obj
    self.member_obj = member_obj
    # 親クラスのdispatchメソッドを呼び出し
    return super().dispatch(request, *args, **kwargs)
  

  # コンテキストデータを設定
  def get_context_data(self, **kwargs):
    context = super().get_context_data(**kwargs)
    request = self.request

    # 設定データ取得
    last_record = administrator_data.objects.order_by("id").last()
    if last_record is None:
      # レコードが1件もない場合、menu_rowフィールドだけに値を設定したインスタンスを作成
      default_data = administrator_data(menu_row=20)
    else:
      default_data = last_record

    # 問い合わせ担当者従業員番号がログイン者の従業員番号と一致している場合、ポップアップ表示設定
    pop_up_display = default_data.administrator_employee_no1 == str(request.session['login_No']) or \
                    default_data.administrator_employee_no2 == str(request.session['login_No']) or \
                    default_data.administrator_employee_no3 == str(request.session['login_No'])

    context.update({
      'title': 'MENU',
      'data': self.member_obj,
      'pop_up_display': pop_up_display,
      'pop_up': default_data,
      })
    
    return context


  # POSTリクエストの処理
  def post(self, request):
    # セッションを削除
    request.session.flush()
    # ログインページに飛ぶ
    return redirect('/login')





#--------------------------------------------------------------------------------------------------------





# 工数メイン画面定義
class KosuMainView(TemplateView):
  # テンプレート定義
  template_name = 'kosu/kosu_main.html'
    

  # リクエストを処理するメソッドをオーバーライド
  def dispatch(self, request, *args, **kwargs):
    # 人員情報取得
    member_obj = get_member(request)
    # 人員情報なしor未ログインの場合ログイン画面へ
    if isinstance(member_obj, HttpResponseRedirect):
      return member_obj
    self.member_obj = member_obj
    # 親クラスのdispatchメソッドを呼び出し
    return super().dispatch(request, *args, **kwargs)


  # コンテキストデータを設定
  def get_context_data(self, **kwargs):
    context = super().get_context_data(**kwargs)

    # HTMLに渡す辞書
    context.update({
        'title': '工数MENU',
        'data': self.member_obj,
    })

    return context





#--------------------------------------------------------------------------------------------------------





# 工数区分定義メイン画面定義
class DefMainView(TemplateView):
  # テンプレート定義
  template_name = 'kosu/def_main.html'
    

  # リクエストを処理するメソッドをオーバーライド
  def dispatch(self, request, *args, **kwargs):
    # 人員情報取得
    member_obj = get_member(request)
    # 人員情報なしor未ログインの場合ログイン画面へ
    if isinstance(member_obj, HttpResponseRedirect):
      return member_obj
    self.member_obj = member_obj
    # 親クラスのdispatchメソッドを呼び出し
    return super().dispatch(request, *args, **kwargs)


  # コンテキストデータを設定
  def get_context_data(self, **kwargs):
    context = super().get_context_data(**kwargs)

    # HTMLに渡す辞書
    context.update({
        'title': '工数区分定義MENU',
        'data': self.member_obj,
    })

    return context





#--------------------------------------------------------------------------------------------------------





# 人員メイン画面定義
class MemberMainView(TemplateView):
  # テンプレート定義
  template_name = 'kosu/member_main.html'
    

  # リクエストを処理するメソッドをオーバーライド
  def dispatch(self, request, *args, **kwargs):
    # 人員情報取得
    member_obj = get_member(request)
    # 人員情報なしor未ログインの場合ログイン画面へ
    if isinstance(member_obj, HttpResponseRedirect):
      return member_obj
    self.member_obj = member_obj

    # ログイン者に権限がなければメインページに戻る
    if member_obj.authority != True:
      return redirect('/')

    # 親クラスのdispatchメソッドを呼び出し
    return super().dispatch(request, *args, **kwargs)
    


  # コンテキストデータを設定
  def get_context_data(self, **kwargs):
    context = super().get_context_data(**kwargs)

    # HTMLに渡す辞書
    context.update({
        'title': '人員MENU',
        'data': self.member_obj,
    })

    return context





#--------------------------------------------------------------------------------------------------------





# 班員メイン画面定義
class TeamMainView(TemplateView):
  # テンプレート定義
  template_name = 'kosu/team_main.html'


  # リクエストを処理するメソッドをオーバーライド
  def dispatch(self, request, *args, **kwargs):
    # 人員情報取得
    member_obj = get_member(request)
    # 人員情報なしor未ログインの場合ログイン画面へ
    if isinstance(member_obj, HttpResponseRedirect):
      return member_obj
    self.member_obj = member_obj

    # ログイン者に権限がなければメインページに戻る
    if member_obj.authority != True:
      return redirect('/')

    # 親クラスのdispatchメソッドを呼び出し
    return super().dispatch(request, *args, **kwargs)


  # コンテキストデータを設定
  def get_context_data(self, **kwargs):
    context = super().get_context_data(**kwargs)

    request = self.request

    # 今日の日時を変数に格納
    today = datetime.date.today()

    # フォロー表示変数定義
    follow_display = False

    # 日付リスト初期状態定義
    day_list = [today - datetime.timedelta(days=d) for d in range(1, 8)]

    # ログイン者の班員登録あるか確認
    team_filter = team_member.objects.filter(employee_no5=request.session['login_No'])

    # ログイン者の班員登録がある場合の処理
    if team_filter.exists():
      # ログイン者の班員情報取得
      team_get = team_member.objects.get(employee_no5=request.session['login_No'])

      # フォロー表示
      follow_display = team_get.follow

      # 班員フォロー用リスト作成
      team_list = []
      for t in range(1, 16):
        member_no = eval(f'team_get.member{t}')
        if member_no:
          member_name_filter = member.objects.filter(employee_no=member_no)
          if member_name_filter.exists():
            member_name_get = member_name_filter.first()
            for ind, dd in enumerate(day_list):
              kosu_filter = Business_Time_graph.objects.filter(
                employee_no3=member_no,
                work_day2=dd
                )
              if kosu_filter.exists():
                kosu_get = kosu_filter.first()
                if kosu_get.tyoku2 != '3' and ind == 0 and not kosu_get.judgement:
                  team_list.append(f'{member_name_get.name}氏の工数未入力があります。')
                  break
                if not kosu_get.judgement and ind != 0:
                  team_list.append(f'{member_name_get.name}氏の工数未入力があります。')
                  break
              else:
                team_list.append(f'{member_name_get.name}氏の工数未入力があります。')
                break
    else:
      # ログイン者の班員登録がない場合の処理
      team_list = []

    # コンテキストにデータを格納
    context.update({
      'title': '班員MENU',
      'data': self.member_obj,
      'team': team_list,
      'follow_display': follow_display,
      })

    return context





#--------------------------------------------------------------------------------------------------------





# 問い合わせメイン画面定義
class InquirMainView(TemplateView):
  # テンプレート定義
  template_name = 'kosu/inquiry_main.html'


  # リクエストを処理するメソッドをオーバーライド
  def dispatch(self, request, *args, **kwargs):
    # 人員情報取得
    member_obj = get_member(request)
    # 人員情報なしor未ログインの場合ログイン画面へ
    if isinstance(member_obj, HttpResponseRedirect):
      return member_obj
    self.member_obj = member_obj

    # ログイン者に権限がなければメインページに戻る
    if member_obj.authority == False:
      return redirect('/')

    # 親クラスのdispatchメソッドを呼び出し
    return super().dispatch(request, *args, **kwargs)
    


  # コンテキストデータを設定
  def get_context_data(self, **kwargs):
    context = super().get_context_data(**kwargs)

    # HTMLに渡す辞書
    context.update({
        'title': '問い合わせMENU',
        'data': self.member_obj,
    })

    return context





#--------------------------------------------------------------------------------------------------------





# 管理者画面定義
class AdministratorMenuView(FormView):
  # テンプレート、フォーム、リダイレクト先定義
  template_name = 'kosu/administrator_menu.html'
  form_class = administrator_data_Form
  success_url = reverse_lazy('administrator')


  # リクエストを処理するメソッドをオーバーライド
  def dispatch(self, request, *args, **kwargs):
    # 人員情報取得
    member_obj = get_member(request)
    # 人員情報なしor未ログインの場合ログイン画面へ
    if isinstance(member_obj, HttpResponseRedirect):
      return member_obj
    self.member_obj = member_obj

    # ログイン者に権限がなければメインページに戻る
    if member_obj.administrator != True:
      return redirect('/')

    # 親クラスのdispatchメソッドを呼び出し
    return super().dispatch(request, *args, **kwargs)


  # 初期データを設定
  def get_initial(self):
    # 設定データ取得
    last_record = administrator_data.objects.order_by("id").last()
    if last_record is None:
      default_data = administrator_data(menu_row=20)
    else:
      default_data = last_record

    # 初期値を設定
    return {
      'menu_row': default_data.menu_row,
      'administrator_employee_no1': default_data.administrator_employee_no1,
      'administrator_employee_no2': default_data.administrator_employee_no2,
      'administrator_employee_no3': default_data.administrator_employee_no3,
      }


  # コンテキストデータを取得するメソッドをオーバーライド
  def get_context_data(self, **kwargs):
    context = super().get_context_data(**kwargs)
    context['title'] = '管理者MENU'
    context['load_form'] = uploadForm()
    return context


  # フォームが有効な場合に呼び出されるメソッドをオーバーライド
  def form_valid(self, form):
    request = self.request

    # 入力内容記録
    edit_comment = f"表示件数:{request.POST['menu_row']}" + '\n' + \
                  f"問い合わせ担当者1:{request.POST['administrator_employee_no1']}" + '\n' + \
                  f"問い合わせ担当者2:{request.POST['administrator_employee_no2']}" + '\n' + \
                  f"問い合わせ担当者3:{request.POST['administrator_employee_no3']}"

    # 設定更新時の処理
    if 'registration' in request.POST:
      # 一覧表示項目数を検証
      menu_row=request.POST.get('menu_row')
      if not menu_row.isdigit() or float(menu_row) <= 0 or math.floor(float(menu_row)) != float(menu_row):
        history_record('管理者MENU', 'administrator_data', 'ERROR029', edit_comment, request)
        messages.error(request, '一覧表示項目数は自然数で入力して下さい。ERROR029')
        return redirect(to='/administrator')

      if has_non_halfwidth_characters(menu_row):
        history_record('管理者MENU', 'administrator_data', 'ERROR056', edit_comment, request)
        messages.error(request, '一覧表示項目数は半角で入力して下さい。ERROR056')
        return redirect(to='/administrator')

      # 問い合わせ担当者従業員番号を検証
      for i in range(1, 4):
        employee_no = request.POST.get(f'administrator_employee_no{i}')
        if employee_no:
          if not employee_no.isdigit() or float(employee_no) <= 0 or math.floor(float(employee_no)) != float(employee_no):
            history_record('管理者MENU', 'administrator_data', 'ERROR051', edit_comment, request)
            messages.error(request, '問い合わせ担当者従業員番号は自然数で入力して下さい。ERROR051')
            return redirect(to='/administrator')

          if has_non_halfwidth_characters(employee_no):
            history_record('管理者MENU', 'administrator_data', 'ERROR057', edit_comment, request)
            messages.error(request, '問い合わせ担当者従業員番号は半角で入力して下さい。ERROR057')
            return redirect(to='/administrator')

          if not member.objects.filter(employee_no=employee_no).exists():
            history_record('管理者MENU', 'administrator_data', 'ERROR058', edit_comment, request)
            messages.error(request, '入力された問い合わせ担当者従業員番号は登録されていません。ERROR058')
            return redirect(to='/administrator')

      # データ更新処理
      last_record = administrator_data.objects.order_by("id").last()
      record_id = last_record.id if last_record else None
      administrator_data.objects.update_or_create(
        id=record_id,
        defaults={
          'menu_row': menu_row,
          'administrator_employee_no1': request.POST.get('administrator_employee_no1'),
          'administrator_employee_no2': request.POST.get('administrator_employee_no2'),
          'administrator_employee_no3': request.POST.get('administrator_employee_no3'),
          })
      history_record('管理者MENU', 'administrator_data', 'OK', edit_comment, request)

    # リダイレクト処理を親クラスへ送る
    return super().form_valid(form)


  # フォームバリデーションが失敗した際の処理
  def form_invalid(self, form):
    request = self.request
    messages.error(request, f'バリテーションエラーが発生しました。IT担当者に連絡してください。{form.errors} ERROR054')
    return redirect(to='/administrator')





#--------------------------------------------------------------------------------------------------------





# ヘルプ画面定義
class HelpView(FormView):
  # テンプレート、フォーム定義
  template_name = 'kosu/help.html'
  form_class = uploadForm


  # コンテキストデータを取得するメソッドをオーバーライド
  def get_context_data(self, **kwargs):
    context = super().get_context_data(**kwargs)
    context['title'] = 'ヘルプ'
    context['display'] = False
    return context


  # フォームが有効な場合の処理
  def form_valid(self, form):
    # ルートディレクトリを取得
    BASE_DIR = Path(__file__).resolve().parent.parent
    # 環境変数ファイルを読み込む
    env = environ.Env()
    env.read_env(os.path.join(BASE_DIR, '.env'))

    # ファイルロード欄非表示
    display = False

    # 復帰パスワード入力時処理
    if 'help_button' in self.request.POST:
      # パスワード判定
      if self.request.POST['help_path'] == env('HELP_PATH'):
        display = True
      else:
        display = False
        messages.error(self.request, 'パスワードが違います。ERROR081')

    # データ読み込み時の処理
    elif 'data_load' in self.request.POST:
      # ファイルロード欄表示
      display = True

      # 人員ファイル未選択時、リダイレクト
      if 'member_file' not in self.request.FILES:
        messages.error(self.request, '人員ファイルが選択されていません。ERROR082')
        return redirect(to='/help')
      # 工数区分定義ファイル未選択時、リダイレクト
      if 'def_file' not in self.request.FILES:
        messages.error(self.request, '工数区分定義ファイルが選択されていません。ERROR083')
        return redirect(to='/help')

      # 人員ファイル定義
      member_file = self.request.FILES['member_file']
      # 一時的なファイルをサーバー上に作成
      with open('member_file_path.xlsx', 'wb+') as destination:
        # アップロードしたファイルを一時ファイルに書き込み
        for chunk in member_file.chunks():
          destination.write(chunk)

      # 一時ファイルを開く
      wb = openpyxl.load_workbook('member_file_path.xlsx')
      ws = wb.worksheets[0]
      # レコード数取得
      data_num = ws.max_row

      # 人員データにレコードがある場合の処理
      if member.objects.exists():
        # Excelからデータを読み込むループ
        for i in range(1, data_num):
          # 人員データに指定従業員番号があるか確認
          employee_no = ws.cell(row=i + 1, column=1).value
          member_obj_filter = member.objects.filter(employee_no=employee_no)

          # 人員データに指定従業員番号がある場合の処理
          if member_obj_filter.exists():
            # 指定人員データを消す
            member_obj_get = member_obj_filter.get()
            member_obj_get.delete()

          # Excelからデータを読み込み
          new_data = member(
            employee_no=employee_no,
            name=ws.cell(row=i + 1, column=2).value,
            shop=ws.cell(row=i + 1, column=3).value,
            authority=ws.cell(row=i + 1, column=4).value,
            administrator=ws.cell(row=i + 1, column=5).value,
            break_time1=ws.cell(row=i + 1, column=6).value,
            break_time1_over1=ws.cell(row=i + 1, column=7).value,
            break_time1_over2=ws.cell(row=i + 1, column=8).value,
            break_time1_over3=ws.cell(row=i + 1, column=9).value,
            break_time2 = ws.cell(row=i + 1, column=10).value,
            break_time2_over1 = ws.cell(row=i + 1, column=11).value,
            break_time2_over2 = ws.cell(row=i + 1, column=12).value,
            break_time2_over3 = ws.cell(row=i + 1, column=13).value,
            break_time3 = ws.cell(row=i + 1, column=14).value,
            break_time3_over1 = ws.cell(row=i + 1, column=15).value,
            break_time3_over2 = ws.cell(row=i + 1, column=16).value,
            break_time3_over3 = ws.cell(row=i + 1, column=17).value,
            break_time4 = ws.cell(row=i + 1, column=18).value,
            break_time4_over1 = ws.cell(row=i + 1, column=19).value,
            break_time4_over2 = ws.cell(row=i + 1, column=20).value,
            break_time4_over3 = ws.cell(row=i + 1, column=21).value,
            break_time5 = ws.cell(row=i + 1, column=22).value,
            break_time5_over1 = ws.cell(row=i + 1, column=23).value,
            break_time5_over2 = ws.cell(row=i + 1, column=24).value,
            break_time5_over3 = ws.cell(row=i + 1, column=25).value,
            break_time6 = ws.cell(row=i + 1, column=26).value,
            break_time6_over1 = ws.cell(row=i + 1, column=27).value,
            break_time6_over2 = ws.cell(row=i + 1, column=28).value,
            break_time6_over3 = ws.cell(row=i + 1, column=29).value,
            pop_up1 = ws.cell(row=i + 1, column=30).value,
            pop_up_id1 = ws.cell(row=i + 1, column=31).value,
            pop_up2 = ws.cell(row=i + 1, column=32).value,
            pop_up_id2 = ws.cell(row=i + 1, column=33).value,
            pop_up3 = ws.cell(row=i + 1, column=34).value,
            pop_up_id3 = ws.cell(row=i + 1, column=35).value,
            pop_up4 = ws.cell(row=i + 1, column=36).value,
            pop_up_id4 = ws.cell(row=i + 1, column=37).value,
            pop_up5 = ws.cell(row=i + 1, column=38).value,
            pop_up_id5 = ws.cell(row=i + 1, column=39).value,
            break_check = ws.cell(row=i + 1, column=40).value,
            def_prediction = ws.cell(row=i + 1, column=41).value
            )
          new_data.save()

      # 人員データにレコードがない場合の処理
      else:
        # Excelからデータを読み込むループ
        for i in range(1, data_num):
          # Excelからデータを読み込み
          new_data = member(
            employee_no=ws.cell(row=i + 1, column=1).value,
            name=ws.cell(row=i + 1, column=2).value,
            shop=ws.cell(row=i + 1, column=3).value,
            authority=ws.cell(row=i + 1, column=4).value,
            administrator=ws.cell(row=i + 1, column=5).value,
            break_time1=ws.cell(row=i + 1, column=6).value,
            break_time1_over1=ws.cell(row=i + 1, column=7).value,
            break_time1_over2=ws.cell(row=i + 1, column=8).value,
            break_time1_over3=ws.cell(row=i + 1, column=9).value,
            break_time2 = ws.cell(row=i + 1, column=10).value,
            break_time2_over1 = ws.cell(row=i + 1, column=11).value,
            break_time2_over2 = ws.cell(row=i + 1, column=12).value,
            break_time2_over3 = ws.cell(row=i + 1, column=13).value,
            break_time3 = ws.cell(row=i + 1, column=14).value,
            break_time3_over1 = ws.cell(row=i + 1, column=15).value,
            break_time3_over2 = ws.cell(row=i + 1, column=16).value,
            break_time3_over3 = ws.cell(row=i + 1, column=17).value,
            break_time4 = ws.cell(row=i + 1, column=18).value,
            break_time4_over1 = ws.cell(row=i + 1, column=19).value,
            break_time4_over2 = ws.cell(row=i + 1, column=20).value,
            break_time4_over3 = ws.cell(row=i + 1, column=21).value,
            pop_up1 = ws.cell(row=i + 1, column=22).value,
            pop_up_id1 = ws.cell(row=i + 1, column=23).value,
            pop_up2 = ws.cell(row=i + 1, column=24).value,
            pop_up_id2 = ws.cell(row=i + 1, column=25).value,
            pop_up3 = ws.cell(row=i + 1, column=26).value,
            pop_up_id3 = ws.cell(row=i + 1, column=27).value,
            pop_up4 = ws.cell(row=i + 1, column=28).value,
            pop_up_id4 = ws.cell(row=i + 1, column=29).value,
            pop_up5 = ws.cell(row=i + 1, column=30).value,
            pop_up_id5 = ws.cell(row=i + 1, column=31).value,
            break_check = ws.cell(row=i + 1, column=32).value,
            def_prediction=ws.cell(row=i + 1, column=33).value,
            break_time5 = ws.cell(row=i + 1, column=22).value,
            break_time5_over1 = ws.cell(row=i + 1, column=23).value,
            break_time5_over2 = ws.cell(row=i + 1, column=24).value,
            break_time5_over3 = ws.cell(row=i + 1, column=25).value,
            break_time6 = ws.cell(row=i + 1, column=26).value,
            break_time6_over1 = ws.cell(row=i + 1, column=27).value,
            break_time6_over2 = ws.cell(row=i + 1, column=28).value,
            break_time6_over3 = ws.cell(row=i + 1, column=29).value,
            pop_up1 = ws.cell(row=i + 1, column=30).value,
            pop_up_id1 = ws.cell(row=i + 1, column=31).value,
            pop_up2 = ws.cell(row=i + 1, column=32).value,
            pop_up_id2 = ws.cell(row=i + 1, column=33).value,
            pop_up3 = ws.cell(row=i + 1, column=34).value,
            pop_up_id3 = ws.cell(row=i + 1, column=35).value,
            pop_up4 = ws.cell(row=i + 1, column=36).value,
            pop_up_id4 = ws.cell(row=i + 1, column=37).value,
            pop_up5 = ws.cell(row=i + 1, column=38).value,
            pop_up_id5 = ws.cell(row=i + 1, column=39).value,
            break_check = ws.cell(row=i + 1, column=40).value,
            def_prediction = ws.cell(row=i + 1, column=41).value
            )
          new_data.save()

      # 一時ファイル削除
      os.remove('member_file_path.xlsx')

      # 工数定義区分ファイル定義
      def_file = self.request.FILES['def_file']
      # 一時的なファイルをサーバー上に作成
      with open('def_file_path.xlsx', 'wb+') as destination:
        # アップロードしたファイルを一時ファイルに書き込み
        for chunk in def_file.chunks():
          destination.write(chunk)

      # 一時ファイルを開く
      wb1 = openpyxl.load_workbook('def_file_path.xlsx')
      ws1 = wb1.worksheets[0]
      # レコード数取得
      data_num = ws1.max_row

      # 工数区分定義データにレコードがある場合の処理
      if kosu_division.objects.exists():
        # Excelからデータを読み込むループ
        for i in range(1, data_num):
          # 工数区分定義データに指定従業員番号があるか確認
          kosu_name = ws1.cell(row=i + 1, column=1).value
          def_obj_filter = kosu_division.objects.filter(kosu_name=kosu_name)

          # 工数区分定義データに指定工数区分名がある場合の処理
          if def_obj_filter.exists():
            # 指定工数区分定義データを消す
            def_obj_get = def_obj_filter.get()
            def_obj_get.delete()

          # Excelからデータを読み込み
          def_data = {'kosu_name': kosu_name}
          for n in range(1, 51):
            def_data[f'kosu_title_{n}'] = ws1.cell(row=i + 1, column=n * 3 - 1).value
            def_data[f'kosu_division_1_{n}'] = ws1.cell(row=i + 1, column=n * 3).value
            def_data[f'kosu_division_2_{n}'] = ws1.cell(row=i + 1, column=n * 3 + 1).value
          new_data1 = kosu_division(**def_data)
          new_data1.save()

      # 工数区分定義データにレコードがない場合の処理
      else:
        # Excelからデータを読み込むループ
        for i in range(1, data_num):
          # Excelからデータを読み込み
          def_data = {'kosu_name': ws1.cell(row=i + 1, column=1).value}
          for n in range(1, 51):
            def_data[f'kosu_title_{n}'] = ws1.cell(row=i + 1, column=n * 3 - 1).value
            def_data[f'kosu_division_1_{n}'] = ws1.cell(row=i + 1, column=n * 3).value
            def_data[f'kosu_division_2_{n}'] = ws1.cell(row=i + 1, column=n * 3 + 1).value
          new_data1 = kosu_division(**def_data)
          new_data1.save()

      # 一時ファイル削除
      os.remove('def_file_path.xlsx')

      return redirect(to='/login')

    # コンテキストデータ更新
    context = self.get_context_data()
    context['display'] = display
    return render(self.request, self.template_name, context)


  # フォームバリデーションが失敗した際の処理
  def form_invalid(self, form):
    request = self.request
    messages.error(request, f'バリテーションエラーが発生しました。IT担当者に連絡してください。{form.errors} ERROR055')
    return redirect(to='/administrator')





#--------------------------------------------------------------------------------------------------------





# データ変更情報一覧表示画面定義
class HistoryList(ListView):
  model = Operation_history
  template_name = 'kosu/history_list.html'
  context_object_name = 'data'
  paginate_by = 20


  # 画面処理前の初期設定
  def dispatch(self, request, *args, **kwargs):
    # 人員情報取得
    member_obj = get_member(request)
    # 人員情報なしor未ログインの場合ログイン画面へ
    if isinstance(member_obj, HttpResponseRedirect):
      return member_obj
    self.member_obj = member_obj
    
    # 権限がないユーザーの場合ログイン画面へ
    if not self.member_obj.administrator:
      return redirect('/')
    return super().dispatch(request, *args, **kwargs)


  # POST時の処理をオーバーライド
  def post(self, request, *args, **kwargs):
    delete_day = request.POST.get('delete_day')
    if delete_day:
      try:
        # POSTされた日付を時間ありに変更
        naive_datetime = datetime.datetime.strptime(delete_day, '%Y-%m-%d')
        aware_datetime = make_aware(naive_datetime)
        # レコード削除
        Operation_history.objects.filter(created_at__lt=aware_datetime).delete()
      except ValueError:
        pass

      # フィルタリング条件をセッションに保存
    request.session['filter_day'] = request.POST.get('day', '')
    request.session['filter_name_list'] = request.POST.get('name_list', '')
    request.session['filter_model_list'] = request.POST.get('model_list', '')
    request.session['filter_page_list'] = request.POST.get('page_list', '')

    return redirect(reverse_lazy('history_list', args = [1]))


  # 1ページのレコード表示数オーバーライド
  def get_paginate_by(self, queryset):
    # 設定データ取得
    last_record = administrator_data.objects.order_by("id").last()
    # レコードが1件もない場合20件設定
    if last_record is None:
      return 20
    return last_record.menu_row


  # フィルタリングしたデータ取得
  def get_queryset(self):
    # セッションからフィルタリング条件を取得
    day = self.request.session.get('filter_day', '')
    name_list = self.request.session.get('filter_name_list', '')
    model_list = self.request.session.get('filter_model_list', '')
    page_list = self.request.session.get('filter_page_list', '')

    # POST時のフィルタ処理
    if day or name_list or model_list or page_list:
      queryset = Operation_history.objects.filter(
        created_at__contains=day,
        employee_no4__contains=name_list,
        operation_models__contains=model_list,
        post_page__contains=page_list
      ).order_by('-created_at')
    # GET時のフィルタ処理
    else:
      queryset = Operation_history.objects.all().order_by('-created_at')

    return queryset


  # フォームの状態定義
  def get_form(self):
    # 編集履歴にある要素リスト作成
    employee_no_list = Operation_history.objects.values_list('employee_no4', flat=True).order_by('employee_no4').distinct()
    model_edit_list = Operation_history.objects.values_list('operation_models', flat=True).order_by('operation_models').distinct()
    page_edit_list = Operation_history.objects.values_list('post_page', flat=True).order_by('post_page').distinct()

    # 選択肢リスト定義
    name_list = [['', '']] + [
      [No, member.objects.get(employee_no=No)]
      for No in employee_no_list
      if member.objects.filter(employee_no=No).exists()
    ]
    model_list = [['', '']] + [[model, model] for model in model_edit_list]
    page_list = [['', '']] + [[page, page] for page in page_edit_list]

    # フォームを作成し選択肢を設定
    if self.request.method == 'POST':
      form = history_findForm(self.request.POST)
    else:
      form = history_findForm()

    form.fields['name_list'].choices = name_list
    form.fields['model_list'].choices = model_list
    form.fields['page_list'].choices = page_list
    return form


  # コンテキスト定義
  def get_context_data(self, **kwargs):
    context = super().get_context_data(**kwargs)
    context['title'] = '編集履歴一覧'
    context['pk'] = self.kwargs.get('pk')
    context['form'] = self.get_form()
    return context





#--------------------------------------------------------------------------------------------------------





# データ操作履歴詳細画面定義
class HistoryDelete(DeleteView):
  # モデル、テンプレート、リダイレクト先などを指定
  model = Operation_history
  template_name = 'kosu/history_delete.html'
  success_url = reverse_lazy('history_list', args = [1])


  # リクエストを処理するメソッドをオーバーライド
  def dispatch(self, request, *args, **kwargs):
    # 人員情報取得
    member_obj = get_member(request)
    # 人員情報なしor未ログインの場合ログイン画面へ
    if isinstance(member_obj, HttpResponseRedirect):
      return member_obj
    self.member_obj = member_obj
    # 親クラスのdispatchメソッドを呼び出し
    return super().dispatch(request, *args, **kwargs)


  # コンテキストデータを取得するメソッドをオーバーライド
  def get_context_data(self, **kwargs):
    context = super().get_context_data(**kwargs)
    obj_get = self.get_object()
    context['title'] = '工数データ削除'
    context['id'] = self.object.id
    context['obj'] = obj_get
    return context





#--------------------------------------------------------------------------------------------------------





# 専用ロガー取得('views_logger'という名前のカスタムロガーを使用しログメッセージ記録)
views_logger = logging.getLogger('views_logger')

# 標準出力をロガーにリダイレクト(sys.stdoutをカスタムロガーへリダイレクト)
# print関数の出力をログとして記録
class PrintToLogger:
  def write(self, message):
    # 受け取った標準出力のメッセージ処理(空白および改行のみのメッセージは無視)
    if message.strip():  
      views_logger.info(message)

  def flush(self):
    # 処理不要のため空処理
    pass

# 標準出力をロガーへリダイレクト(sys.stdoutをPrintToLoggerのインスタンスに設定)
sys.stdout = PrintToLogger()

# 記録されたログを画面上に表示
def get_logs(request):
  # 'web_console.log' ファイルを読み込み専用モードで開く
  with open('web_console.log', 'r') as log_file:
    # ファイルの内容を行ごとにリストとして格納
    logs = log_file.readlines()
  # JSONレスポンスとして、ログを返す
  return JsonResponse({'logs': logs})





#--------------------------------------------------------------------------------------------------------

