import os
from django.conf import settings
import openpyxl
import pandas as pd
import tempfile
import datetime
from .models import Business_Time_graph, kosu_division, member, team_member, inquiry_data
from .utils.kosu_utils import kosu_division_dictionary





#--------------------------------------------------------------------------------------------------------





# 工数データバックアップ非同期処理
def generate_kosu_backup(data_day, data_day2):
  # 新しいExcelブック作成
  wb = openpyxl.Workbook()
  ws = wb.active

  # ヘッダー作成
  headers = [
    '従業員番号', '氏名', '工数区分定義Ver', '就業日', '直',
    '作業内容', '作業詳細', '残業時間', '昼休憩時間',
    '残業休憩時間1', '残業休憩時間2', '残業休憩時間3',
    '就業形態', '工数入力OK_NG', '休憩変更チェック',
    ]
  ws.append(headers)

  # 工数データ取得
  kosu_data = Business_Time_graph.objects.filter(
    work_day2__gte=data_day, work_day2__lte=data_day2
    )

  # データ書き込み
  for item in kosu_data:
    row = [
      item.employee_no3, str(item.name), item.def_ver2, item.work_day2,
      item.tyoku2, item.time_work, item.detail_work, item.over_time,
      item.breaktime, item.breaktime_over1, item.breaktime_over2,
      item.breaktime_over3, item.work_time, item.judgement,
      item.break_change,
      ]
    ws.append(row)

  # 保存先のディレクトリ確認・作成
  media_dir = settings.MEDIA_ROOT
  if not os.path.exists(media_dir):
    os.makedirs(media_dir)

  # ファイル名作成と保存
  filename = f'工数データバックアップ_{data_day}_{data_day2}.xlsx'
  filepath = os.path.join(media_dir, filename)
  wb.save(filepath)

  # ファイルパスを返却
  return filepath





#--------------------------------------------------------------------------------------------------------





# 工数区分定義予測データ出力非同期処理
def generate_prediction(data_day, data_day2):
  # 期間内の工数データ取得
  kosu_filter = Business_Time_graph.objects.filter(
    work_day2__gte=data_day, work_day2__lte=data_day2
    )

  # Excelに書き出すためのデータを準備
  data = []
  # 取得した工数データを処理
  for kosu in kosu_filter:
    # 作業内容をリストに解凍
    kosu_list = list(kosu.time_work)
    # 作業詳細をリストに解凍
    detail_list = kosu.detail_work.split('$')
    # 工数定義区分取得
    def_filter = kosu_division.objects.filter(kosu_name=kosu.def_ver2)

    # リストの長さ取得
    max_length = max(len(kosu_list), len(detail_list))
    # 1要素ごとにExcelに書き込み
    for i in range(max_length):
      # 工数定義区分がある場合の処理
      if def_filter.exists():
        # 工数定義区分リスト作成
        choices_list, def_n = kosu_division_dictionary(kosu.def_ver2)

        # 作業内容を工数区分定義に変換
        for k in choices_list:
          # 工数区分定義の記号と作業内容が同じ場合の処理
          if k[0] == kosu_list[i]:
            # 作業詳細が空欄でない場合の処理
            if detail_list[i] != '':
              # 作業内容と作業詳細をセットで定義
              row = [
                  k[1] if i < len(kosu_list) else '',
                  detail_list[i] if i < len(detail_list) else '',
              ]
              # 作業内容と作業詳細を書き込み
              data.append(row)
              #ループから抜ける
              break

  # DataFrameに変換
  df = pd.DataFrame(data, columns=['工数定義区分', '作業詳細'])

  # フィルタリング条件に基づいて不要な行を削除
  df = df[~df['工数定義区分'].isin(['', '#', '$'])]
  df = df[df['作業詳細'] != '']

  # 工数定義区分と作業詳細の重複行を削除
  df = df.drop_duplicates(subset=['工数定義区分', '作業詳細'], keep='first')

  # 保存先のディレクトリ確認・作成
  media_dir = settings.MEDIA_ROOT
  if not os.path.exists(media_dir):
    os.makedirs(media_dir)

  # ファイル名作成と保存
  filename = f'工数予測データバックアップ_{data_day}_{data_day2}.xlsx'
  filepath = os.path.join(media_dir, filename)
  
  # Excelファイルをメモリ上で作成して保存
  with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
      df.to_excel(writer, index=False)

  # ファイルパスを返却
  return filepath





#--------------------------------------------------------------------------------------------------------





# 工数データ削除非同期処理
def delete_kosu_data(data_day, data_day2):
  # 工数データ取得
  kosu_obj = Business_Time_graph.objects.filter(work_day2__gte=data_day, work_day2__lte=data_day2)
  # 取得した工数データを削除
  kosu_obj.delete()

  return None




#--------------------------------------------------------------------------------------------------------





# 工数データロード非同期処理
def load_kosu_file(file_obj):
  try:
    # 一時ファイルを作成
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
      # ファイルを書き込む
      for chunk in file_obj.chunks():
          temp_file.write(chunk)

      # ファイル名保存
      temp_file_path = temp_file.name

    # ファイルを開く
    wb = openpyxl.load_workbook(temp_file_path)
    ws = wb.worksheets[0]

    # ヘッダー定義
    expected_headers = [
      '従業員番号', '氏名', '工数区分定義Ver', '就業日', '直',
      '作業内容', '作業詳細', '残業時間', '昼休憩時間',
      '残業休憩時間1', '残業休憩時間2', '残業休憩時間3',
      '就業形態', '工数入力OK_NG', '休憩変更チェック',
      ]
    # ファイル内ヘッダー取得
    actual_headers = [ws.cell(1, col).value for col in range(1, len(expected_headers) + 1)]
    # ヘッダーのデータに相違がある場合、一時ファイル削除しエラーを返す
    if actual_headers != expected_headers:
      os.remove(temp_file_path)
      return {'status': 'error', 'message': '無効なファイルフォーマットです。'}, None

    # データ読み込み
    for i in range(2, ws.max_row + 1):
      employee_no = ws.cell(row=i, column=1).value
      work_day2 = ws.cell(row=i, column=4).value

      # もし既に同一データが存在するなら削除
      existing_data = Business_Time_graph.objects.filter(
        employee_no3=employee_no, work_day2=work_day2
        )
      if existing_data.exists():
        existing_data.delete()

      # 新データをインスタンスとして作成してDBに保存
      Business_Time_graph.objects.create(
        employee_no3=employee_no,
        name=member.objects.get(employee_no=employee_no),
        def_ver2=ws.cell(row=i, column=3).value,
        work_day2=work_day2,
        tyoku2=ws.cell(row=i, column=5).value,
        time_work=ws.cell(row=i, column=6).value,
        detail_work=ws.cell(row=i, column=7).value,
        over_time=ws.cell(row=i, column=8).value,
        breaktime=ws.cell(row=i, column=9).value,
        breaktime_over1=ws.cell(row=i, column=10).value,
        breaktime_over2=ws.cell(row=i, column=11).value,
        breaktime_over3=ws.cell(row=i, column=12).value,
        work_time=ws.cell(row=i, column=13).value,
        judgement=ws.cell(row=i, column=14).value,
        break_change=ws.cell(row=i, column=15).value,
        )

    # 一時ファイルを削除
    os.remove(temp_file_path)
    return {'status': 'success'}, None

  except Exception as e:
    # ロード処理ミスした際は一時ファイルがあれば削除しエラーを返す
    if 'temp_file_path' in locals():
      os.remove(temp_file_path)
    return {'status': 'error', 'message': str(e)}, None





#--------------------------------------------------------------------------------------------------------





# 人員データバックアップ非同期処理
def generate_member_backup():
  # 今日の日付取得
  today = datetime.date.today().strftime('%Y%m%d')
  # 新しいExcelブック作成
  wb = openpyxl.Workbook()
  ws = wb.active

  # ヘッダー作成
  headers = [
    '従業員番号', '氏名', 'ショップ', '権限', '管理者', 
    '1直昼休憩時間', '1直残業休憩時間1', '1直残業休憩時間2', '1直残業休憩時間3', 
    '2直昼休憩時間', '2直残業休憩時間1', '2直残業休憩時間2', '2直残業休憩時間3', 
    '3直昼休憩時間', '3直残業休憩時間1', '3直残業休憩時間2', '3直残業休憩時間3', 
    '常昼昼休憩時間', '常昼残業休憩時間1', '常昼残業休憩時間2', '常昼残業休憩時間3',
    'ポップアップ1', 'ポップアップID1', 'ポップアップ2', 'ポップアップID2',
    'ポップアップ3', 'ポップアップID3', 'ポップアップ4', 'ポップアップID4',
    'ポップアップ5', 'ポップアップID6', '休憩エラー有効チェック', '工数定義区分予測無効',
    ]
  ws.append(headers)

  # 人員データ取得
  member_data = member.objects.all()

  # データ書き込み
  for item in member_data:
    row = [
      item.employee_no, item.name, item.shop, item.authority, item.administrator, 
      item.break_time1, item.break_time1_over1, item.break_time1_over2, item.break_time1_over3, 
      item.break_time2, item.break_time2_over1, item.break_time2_over2, item.break_time2_over3, 
      item.break_time3, item.break_time3_over1, item.break_time3_over2, item.break_time3_over3, 
      item.break_time4, item.break_time4_over1, item.break_time4_over2, item.break_time4_over3,
      item.pop_up1, item.pop_up_id1, item.pop_up2, item.pop_up_id2, item.pop_up3, item.pop_up_id3, 
      item.pop_up4, item.pop_up_id4, item.pop_up5, item.pop_up_id5, item.break_check, item.def_prediction, 
      ]
    ws.append(row)

  # 保存先のディレクトリ確認・作成
  media_dir = settings.MEDIA_ROOT
  if not os.path.exists(media_dir):
    os.makedirs(media_dir)

  # ファイル名作成と保存
  filename = f'人員データバックアップ_{today}.xlsx'
  filepath = os.path.join(media_dir, filename)
  wb.save(filepath)

  # ファイルパスを返却
  return filepath





#--------------------------------------------------------------------------------------------------------





# 人員データロード非同期処理
def load_member_file(request, file_obj):
  try:
    # 一時ファイルを作成
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
      # ファイルを書き込む
      for chunk in file_obj.chunks():
          temp_file.write(chunk)

      # ファイル名保存
      temp_file_path = temp_file.name

    # ファイルを開く
    wb = openpyxl.load_workbook(temp_file_path)
    ws = wb.worksheets[0]

    # ヘッダー定義
    expected_headers = [
      '従業員番号', '氏名', 'ショップ', '権限', '管理者', 
      '1直昼休憩時間', '1直残業休憩時間1', '1直残業休憩時間2', '1直残業休憩時間3', 
      '2直昼休憩時間', '2直残業休憩時間1', '2直残業休憩時間2', '2直残業休憩時間3', 
      '3直昼休憩時間', '3直残業休憩時間1', '3直残業休憩時間2', '3直残業休憩時間3', 
      '常昼昼休憩時間', '常昼残業休憩時間1', '常昼残業休憩時間2', '常昼残業休憩時間3',
      'ポップアップ1', 'ポップアップID1', 'ポップアップ2', 'ポップアップID2',
      'ポップアップ3', 'ポップアップID3', 'ポップアップ4', 'ポップアップID4',
      'ポップアップ5', 'ポップアップID6', '休憩エラー有効チェック', '工数定義区分予測無効',
      ]

    # ファイル内ヘッダー取得
    actual_headers = [ws.cell(1, col).value for col in range(1, len(expected_headers) + 1)]
    # ヘッダーのデータに相違がある場合、一時ファイル削除しエラーを返す
    if actual_headers != expected_headers:
      os.remove(temp_file_path)
      return {'status': 'error', 'message': '無効なファイルフォーマットです。'}, None

    # データ読み込み
    for i in range(2, ws.max_row + 1):
      # 読み込み予定データと同一の従業員番号のデータが存在するか確認
      member_data_filter = member.objects.filter(employee_no = ws.cell(row = i, column = 1).value)
      # 上書きチェックONの場合の処理
      if ('overwrite_check' in request.POST):
        # 新データをインスタンスとして作成してDBに保存
        member.objects.create(
          employee_no=ws.cell(row=i, column=1).value,
          name=ws.cell(row=i, column=2).value,
          shop=ws.cell(row=i, column=3).value,
          authority = ws.cell(row=i, column=4).value,
          administrator = ws.cell(row=i, column=5).value,
          break_time1 = ws.cell(row=i, column=6).value,
          break_time1_over1 = ws.cell(row=i, column=7).value,
          break_time1_over2 = ws.cell(row=i, column=8).value,
          break_time1_over3 = ws.cell(row=i, column=9).value,
          break_time2 = ws.cell(row=i, column=10).value,
          break_time2_over1 = ws.cell(row=i, column=11).value,
          break_time2_over2 = ws.cell(row=i, column=12).value,
          break_time2_over3 = ws.cell(row=i, column=13).value,
          break_time3 = ws.cell(row=i, column=14).value,
          break_time3_over1 = ws.cell(row=i, column=15).value,
          break_time3_over2 = ws.cell(row=i, column=16).value,
          break_time3_over3 = ws.cell(row=i, column=17).value,
          break_time4 = ws.cell(row=i, column=18).value,
          break_time4_over1 = ws.cell(row=i, column=19).value,
          break_time4_over2 = ws.cell(row=i, column=20).value,
          break_time4_over3 = ws.cell(row=i, column=21).value,
          pop_up1 = ws.cell(row=i, column=22).value,
          pop_up_id1 = ws.cell(row=i, column=23).value,
          pop_up2 = ws.cell(row=i, column=24).value,
          pop_up_id2 = ws.cell(row=i, column=25).value,
          pop_up3 = ws.cell(row=i, column=26).value,
          pop_up_id3 = ws.cell(row=i, column=27).value,
          pop_up4 = ws.cell(row=i, column=28).value,
          pop_up_id4 = ws.cell(row=i, column=29).value,
          pop_up5 = ws.cell(row=i, column=30).value,
          pop_up_id5 = ws.cell(row=i, column=31).value,
          break_check = ws.cell(row=i, column=32).value,
          def_prediction = ws.cell(row=i, column=33).value
          )
        
      # 上書きチェックOFFの場合の処理
      else:
        # 読み込み予定データと同一の従業員番号のデータが存在する場合の処理
        if not member_data_filter.exists():
          # 新データをインスタンスとして作成してDBに保存
          member.objects.create(
            employee_no=ws.cell(row=i, column=1).value,
            name=ws.cell(row=i, column=2).value,
            shop=ws.cell(row=i, column=3).value,
            authority = ws.cell(row=i, column=4).value,
            administrator = ws.cell(row=i, column=5).value,
            break_time1 = ws.cell(row=i, column=6).value,
            break_time1_over1 = ws.cell(row=i, column=7).value,
            break_time1_over2 = ws.cell(row=i, column=8).value,
            break_time1_over3 = ws.cell(row=i, column=9).value,
            break_time2 = ws.cell(row=i, column=10).value,
            break_time2_over1 = ws.cell(row=i, column=11).value,
            break_time2_over2 = ws.cell(row=i, column=12).value,
            break_time2_over3 = ws.cell(row=i, column=13).value,
            break_time3 = ws.cell(row=i, column=14).value,
            break_time3_over1 = ws.cell(row=i, column=15).value,
            break_time3_over2 = ws.cell(row=i, column=16).value,
            break_time3_over3 = ws.cell(row=i, column=17).value,
            break_time4 = ws.cell(row=i, column=18).value,
            break_time4_over1 = ws.cell(row=i, column=19).value,
            break_time4_over2 = ws.cell(row=i, column=20).value,
            break_time4_over3 = ws.cell(row=i, column=21).value,
            pop_up1 = ws.cell(row=i, column=22).value,
            pop_up_id1 = ws.cell(row=i, column=23).value,
            pop_up2 = ws.cell(row=i, column=24).value,
            pop_up_id2 = ws.cell(row=i, column=25).value,
            pop_up3 = ws.cell(row=i, column=26).value,
            pop_up_id3 = ws.cell(row=i, column=27).value,
            pop_up4 = ws.cell(row=i, column=28).value,
            pop_up_id4 = ws.cell(row=i, column=29).value,
            pop_up5 = ws.cell(row=i, column=30).value,
            pop_up_id5 = ws.cell(row=i, column=31).value,
            break_check = ws.cell(row=i, column=32).value,
            def_prediction = ws.cell(row=i, column=33).value
            )

    # 一時ファイルを削除
    os.remove(temp_file_path)
    return {'status': 'success'}, None

  except Exception as e:
    # ロード処理ミスした際は一時ファイルがあれば削除しエラーを返す
    if 'temp_file_path' in locals():
      os.remove(temp_file_path)
    return {'status': 'error', 'message': str(e)}, None





#--------------------------------------------------------------------------------------------------------





# 班員データバックアップ非同期処理
def generate_team_backup():
  # 今日の日付取得
  today = datetime.date.today().strftime('%Y%m%d')
  # 新しいExcelブック作成
  wb = openpyxl.Workbook()
  ws = wb.active

  # ヘッダー作成
  headers = [
    '従業員番号', '班員1', '班員2', '班員3', '班員4', 
    '班員5', '班員6', '班員7', '班員8', '班員9', '班員10',
    '班員11', '班員12', '班員13', '班員14', '班員15',
    'フォローON/OFF',
    ]

  ws.append(headers)

  # 班員データ取得
  team_data = team_member.objects.all()

  # データ書き込み
  for item in team_data:
    row = [
      item.employee_no5, item.member1, item.member2, item.member3, 
      item.member4, item.member5, item.member6, item.member7, 
      item.member8, item.member9, item.member10,item.member11, 
      item.member12, item.member13, item.member14, item.member15,
      item.follow, 
      ]

    ws.append(row)

  # 保存先のディレクトリ確認・作成
  media_dir = settings.MEDIA_ROOT
  if not os.path.exists(media_dir):
    os.makedirs(media_dir)

  # ファイル名作成と保存
  filename = f'班員データバックアップ_{today}.xlsx'
  filepath = os.path.join(media_dir, filename)
  wb.save(filepath)

  # ファイルパスを返却
  return filepath





#--------------------------------------------------------------------------------------------------------





# 班員データロード非同期処理
def load_team_file(file_obj):
  try:
    # 一時ファイルを作成
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
      # ファイルを書き込む
      for chunk in file_obj.chunks():
          temp_file.write(chunk)

      # ファイル名保存
      temp_file_path = temp_file.name

    # ファイルを開く
    wb = openpyxl.load_workbook(temp_file_path)
    ws = wb.worksheets[0]

    # ヘッダー定義
    expected_headers = [
      '従業員番号', '班員1', '班員2', '班員3', '班員4', 
      '班員5', '班員6', '班員7', '班員8', '班員9', '班員10',
      '班員11', '班員12', '班員13', '班員14', '班員15',
      'フォローON/OFF',
      ]

    # ファイル内ヘッダー取得
    actual_headers = [ws.cell(1, col).value for col in range(1, len(expected_headers) + 1)]
    # ヘッダーのデータに相違がある場合、一時ファイル削除しエラーを返す
    if actual_headers != expected_headers:
      os.remove(temp_file_path)
      return {'status': 'error', 'message': '無効なファイルフォーマットです。'}, None

    # データ読み込み
    for i in range(2, ws.max_row + 1):
      # 読み込み予定データと同一の従業員番号のデータが存在するか確認
      team_data_filter = team_member.objects.filter(employee_no5=ws.cell(row = i, column = 1).value)
      # 同一従業員番号のデータがあった場合データ削除
      if team_data_filter.exists():
        team_data_filter.delete()

      # 新データをインスタンスとして作成してDBに保存
      team_member.objects.create(
        employee_no5=ws.cell(row=i, column=1).value,
        member1=ws.cell(row=i, column=2).value,
        member2=ws.cell(row=i, column=3).value,
        member3=ws.cell(row=i, column=4).value,
        member4=ws.cell(row=i, column=5).value,
        member5=ws.cell(row=i, column=6).value,
        member6=ws.cell(row=i, column=7).value,
        member7=ws.cell(row=i, column=8).value,
        member8=ws.cell(row=i, column=9).value,
        member9=ws.cell(row=i, column=10).value,
        member10=ws.cell(row=i, column=11).value,
        member11=ws.cell(row=i, column=12).value,
        member12=ws.cell(row=i, column=13).value,
        member13=ws.cell(row=i, column=14).value,
        member14=ws.cell(row=i, column=15).value,
        member15=ws.cell(row=i, column=16).value,
        follow=ws.cell(row=i, column=17).value,
        )

    # 一時ファイルを削除
    os.remove(temp_file_path)
    return {'status': 'success'}, None

  except Exception as e:
    # ロード処理ミスした際は一時ファイルがあれば削除しエラーを返す
    if 'temp_file_path' in locals():
      os.remove(temp_file_path)
    return {'status': 'error', 'message': str(e)}, None





#--------------------------------------------------------------------------------------------------------





# 工数区分定義データバックアップ非同期処理
def generate_def_backup():
  # 今日の日付取得
  today = datetime.date.today().strftime('%Y%m%d')
  # 新しいExcelブック作成
  wb = openpyxl.Workbook()
  ws = wb.active

  # ヘッダー作成
  headers = ['工数区分定義Ver名'] + [item for i in range(1, 51) for item in [f'工数区分名{i}', f'定義{i}', f'作業内容{i}']]

  ws.append(headers)

  # 班員データ取得
  def_data = kosu_division.objects.all()

  # データ書き込み
  for item in def_data:
    row = [item.kosu_name] + \
          [getattr(item, f'kosu_title_{i}') if j == 0 else getattr(item, f'kosu_division_{j}_{i}') 
          for i in range(1, 51) for j in range(0, 3)]
    ws.append(row)

  # 保存先のディレクトリ確認・作成
  media_dir = settings.MEDIA_ROOT
  if not os.path.exists(media_dir):
    os.makedirs(media_dir)

  # ファイル名作成と保存
  filename = f'工数区分定義データバックアップ_{today}.xlsx'
  filepath = os.path.join(media_dir, filename)
  wb.save(filepath)

  # ファイルパスを返却
  return filepath





#--------------------------------------------------------------------------------------------------------





# 工数区分定義データロード非同期処理
def load_def_file(file_obj):
  try:
    # 一時ファイルを作成
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
      # ファイルを書き込む
      for chunk in file_obj.chunks():
          temp_file.write(chunk)

      # ファイル名保存
      temp_file_path = temp_file.name

    # ファイルを開く
    wb = openpyxl.load_workbook(temp_file_path)
    ws = wb.worksheets[0]

    # ヘッダー定義
    expected_headers = ['工数区分定義Ver名'] + [item for i in range(1, 51) for item in [f'工数区分名{i}', f'定義{i}', f'作業内容{i}']]

    # ファイル内ヘッダー取得
    actual_headers = [ws.cell(1, col).value for col in range(1, len(expected_headers) + 1)]
    # ヘッダーのデータに相違がある場合、一時ファイル削除しエラーを返す
    if actual_headers != expected_headers:
      os.remove(temp_file_path)
      return {'status': 'error', 'message': '無効なファイルフォーマットです。'}, None

    # データ読み込み
    for i in range(2, ws.max_row + 1):
      # 読み込み予定データと同一の工数区分定義データが存在するか確認
      def_data_filter = kosu_division.objects.filter(kosu_name=ws.cell(row = i, column = 1).value)
      # 同一工数区分定義データがあった場合データ削除
      if def_data_filter.exists():
        def_data_filter.delete()

      # データ定義
      def_data = {
        'kosu_name': ws.cell(row=i, column=1).value
        }
      for n in range(1, 51):
        def_data[f'kosu_title_{n}'] = ws.cell(row=i, column=n*3-1).value
        def_data[f'kosu_division_1_{n}'] = ws.cell(row=i, column=n*3).value
        def_data[f'kosu_division_2_{n}'] = ws.cell(row=i, column=n*3+1).value

      new_data = kosu_division(**def_data)
      new_data.save()

    # 一時ファイルを削除
    os.remove(temp_file_path)
    return {'status': 'success'}, None

  except Exception as e:
    # ロード処理ミスした際は一時ファイルがあれば削除しエラーを返す
    if 'temp_file_path' in locals():
      os.remove(temp_file_path)
    return {'status': 'error', 'message': str(e)}, None





#--------------------------------------------------------------------------------------------------------





# 問い合わせデータバックアップ非同期処理
def generate_inquiry_backup():
  # 今日の日付取得
  today = datetime.date.today().strftime('%Y%m%d')
  # 新しいExcelブック作成
  wb = openpyxl.Workbook()
  ws = wb.active

  # ヘッダー作成
  headers = [
    '従業員番号', '氏名', '内容選択', '問い合わせ', '回答'
    ]

  ws.append(headers)

  # 問い合わせデータ取得
  inquiry = inquiry_data.objects.all()

  # データ書き込み
  for item in inquiry:
    row = [
        item.employee_no2, 
        str(item.name), 
        item.content_choice, 
        item.inquiry, 
        item.answer,
      ]

    ws.append(row)

  # 保存先のディレクトリ確認・作成
  media_dir = settings.MEDIA_ROOT
  if not os.path.exists(media_dir):
    os.makedirs(media_dir)

  # ファイル名作成と保存
  filename = f'問い合わせデータバックアップ_{today}.xlsx'
  filepath = os.path.join(media_dir, filename)
  wb.save(filepath)

  # ファイルパスを返却
  return filepath





#--------------------------------------------------------------------------------------------------------





# 問い合わせデータロード非同期処理
def load_inquiry_file(file_obj):
  try:
    # 一時ファイルを作成
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
      # ファイルを書き込む
      for chunk in file_obj.chunks():
          temp_file.write(chunk)

      # ファイル名保存
      temp_file_path = temp_file.name

    # ファイルを開く
    wb = openpyxl.load_workbook(temp_file_path)
    ws = wb.worksheets[0]

    # ヘッダー定義
    expected_headers = [
    '従業員番号', '氏名', '内容選択', '問い合わせ', '回答'
      ]

    # ファイル内ヘッダー取得
    actual_headers = [ws.cell(1, col).value for col in range(1, len(expected_headers) + 1)]
    # ヘッダーのデータに相違がある場合、一時ファイル削除しエラーを返す
    if actual_headers != expected_headers:
      os.remove(temp_file_path)
      return {'status': 'error', 'message': '無効なファイルフォーマットです。'}, None

    # データ読み込み
    for i in range(2, ws.max_row + 1):
      # 従業員番号取得
      employee_no = ws.cell(row=i, column=1).value

      # 新データをインスタンスとして作成してDBに保存
      inquiry_data.objects.create(
        employee_no2=employee_no,
        name=member.objects.get(employee_no=employee_no),
        content_choice=ws.cell(row=i, column=3).value,
        inquiry=ws.cell(row=i, column=4).value,
        answer=ws.cell(row=i, column=5).value,
        )

    # 一時ファイルを削除
    os.remove(temp_file_path)
    return {'status': 'success'}, None

  except Exception as e:
    # ロード処理ミスした際は一時ファイルがあれば削除しエラーを返す
    if 'temp_file_path' in locals():
      os.remove(temp_file_path)
    return {'status': 'error', 'message': str(e)}, None





#--------------------------------------------------------------------------------------------------------



